"""
Databricks App Entry Point
Combines React frontend with FastAPI backend following Databricks Apps best practices
Based on: https://www.databricks.com/blog/building-databricks-apps-react-and-mosaic-ai-agents-enterprise-chat-solutions
"""

from fastapi import FastAPI, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from typing import List, Dict, Any, Optional
import os
import requests
import time
import json
import io
import uuid
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================================
# API Application (Backend Logic)
# ============================================================================

api_app = FastAPI(title="Data Profiler API", version="1.0.0")

# CORS middleware for API
api_app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ============================================================================
# Databricks Connection Management
# ============================================================================

def get_oauth_token():
    """Get OAuth token using client credentials (Databricks Apps)"""
    workspace_host = os.getenv("DATABRICKS_SERVER_HOSTNAME", "")
    client_id = os.getenv("DATABRICKS_CLIENT_ID", "")
    client_secret = os.getenv("DATABRICKS_CLIENT_SECRET", "")
    
    url = f"https://{workspace_host}/oidc/v1/token"
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}
    data = {
        'grant_type': 'client_credentials',
        'client_id': client_id,
        'client_secret': client_secret,
        'scope': 'all-apis'
    }
    
    response = requests.post(url, headers=headers, data=data)
    response.raise_for_status()
    return response.json().get('access_token')


def execute_sql(sql: str):
    """Execute SQL using Statements API (same as governance app)"""
    token = get_oauth_token()
    workspace_host = os.getenv("DATABRICKS_SERVER_HOSTNAME", "")
    http_path = os.getenv("DATABRICKS_HTTP_PATH", "")
    warehouse_id = http_path.split('/warehouses/')[-1] if '/warehouses/' in http_path else ""
    
    headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}
    submit_url = f"https://{workspace_host}/api/2.0/sql/statements"
    
    payload = {
        "statement": sql,
        "warehouse_id": warehouse_id,
        "wait_timeout": "30s"
    }
    
    r = requests.post(submit_url, headers=headers, json=payload)
    r.raise_for_status()
    stmt = r.json()
    statement_id = stmt.get("statement_id")
    status = (stmt.get("status") or {}).get("state")
    
    # If already finished
    if status in ("SUCCEEDED", "FAILED", "CANCELED"):
        final = stmt
    else:
        # Poll for completion
        status_url = f"{submit_url}/{statement_id}"
        for _ in range(60):  # up to 60s
            rr = requests.get(status_url, headers=headers)
            rr.raise_for_status()
            final = rr.json()
            state = (final.get("status") or {}).get("state")
            if state in ("SUCCEEDED", "FAILED", "CANCELED"):
                break
            time.sleep(1.0)
    
    state = (final.get("status") or {}).get("state")
    if state != "SUCCEEDED":
        error_msg = (final.get("status") or {}).get("error", {})
        raise RuntimeError(f"SQL failed: state={state}, error={error_msg}")
    
    # Get column names from manifest
    result = final.get("result") or {}
    manifest = result.get("manifest") or {}
    schema = manifest.get("schema") or {}
    columns_meta = schema.get("columns") or []
    
    # Get rows as list of arrays
    data_array = result.get("data_array") or []
    
    # If no column metadata, return raw arrays
    # This is expected for complex queries (CTEs, aggregations, JSON functions)
    # The calling code will manually map array positions to dict keys
    if not columns_meta:
        return data_array
    
    # Extract column names
    column_names = [col.get("name") for col in columns_meta]
    
    # Convert to list of dicts for easier access
    rows_as_dicts = []
    try:
        for row in data_array:
            row_dict = {}
            for i, col_name in enumerate(column_names):
                row_dict[col_name] = row[i] if i < len(row) else None
            rows_as_dicts.append(row_dict)
        return rows_as_dicts
    except Exception as e:
        print(f"Error converting rows to dicts: {e}, returning raw arrays")
        return data_array


# ============================================================================
# Pydantic Models
# ============================================================================

class ProfilingQuery(BaseModel):
    fieldKey: str
    query: str
    description: str


class QueryRequest(BaseModel):
    queries: List[ProfilingQuery]


class CatalogTreeResponse(BaseModel):
    catalogs: List[Dict[str, Any]]


class QueryExecutionResponse(BaseModel):
    results: List[Dict[str, Any]]
    success: bool
    message: Optional[str] = None


# ============================================================================
# API Endpoints
# ============================================================================

@api_app.get("/health")
async def health_check():
    """Health check endpoint for monitoring"""
    try:
        # Test SQL execution
        execute_sql("SELECT 1")
        
        return {
            "status": "healthy",
            "databricks": "connected",
            "app": "data-profiler",
            "version": "1.0.0"
        }
    except Exception as e:
        return {
            "status": "unhealthy",
            "error": str(e)
        }


@api_app.get("/catalogs")
async def get_catalogs():
    """Get list of catalogs only - fast, incremental loading"""
    try:
        catalogs_data = execute_sql("""
            SELECT catalog_name 
            FROM system.information_schema.catalogs 
            WHERE catalog_owner IS NOT NULL 
            ORDER BY catalog_name
        """)
        # Handle both dict (with column names) and array (fallback) formats
        if catalogs_data and isinstance(catalogs_data[0], dict):
            catalogs = [{"name": row["catalog_name"]} for row in catalogs_data]
        else:
            catalogs = [{"name": row[0]} for row in catalogs_data]
        return {"catalogs": catalogs}
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error fetching catalogs: {error_details}")
        raise HTTPException(status_code=500, detail=f"Error fetching catalogs: {str(e)}")


@api_app.get("/schemas")
async def get_schemas(catalog: str):
    """Get schemas for a specific catalog"""
    try:
        schemas_data = execute_sql(f"""
            SELECT schema_name
            FROM system.information_schema.schemata
            WHERE catalog_name = '{catalog}'
                AND schema_name NOT IN ('information_schema', 'system')
            ORDER BY schema_name
        """)
        # Handle both dict and array formats
        if schemas_data and isinstance(schemas_data[0], dict):
            schemas = [{"name": row["schema_name"]} for row in schemas_data]
        else:
            schemas = [{"name": row[0]} for row in schemas_data]
        return {"schemas": schemas}
    except Exception as e:
        import traceback
        print(f"Error fetching schemas: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error fetching schemas: {str(e)}")


@api_app.get("/tables")
async def get_tables(catalog: str, schema: str):
    """Get tables for a specific catalog.schema"""
    try:
        tables_data = execute_sql(f"""
            SELECT table_name
            FROM system.information_schema.tables
            WHERE table_catalog = '{catalog}'
                AND table_schema = '{schema}'
            ORDER BY table_name
        """)
        # Handle both dict and array formats
        if tables_data and isinstance(tables_data[0], dict):
            tables = [{"name": row["table_name"]} for row in tables_data]
        else:
            tables = [{"name": row[0]} for row in tables_data]
        return {"tables": tables}
    except Exception as e:
        import traceback
        print(f"Error fetching tables: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error fetching tables: {str(e)}")


@api_app.get("/columns")
async def get_columns(catalog: str, schema: str, table: str):
    """Get columns for a specific table"""
    try:
        columns_data = execute_sql(f"""
            SELECT column_name, data_type
            FROM system.information_schema.columns
            WHERE table_catalog = '{catalog}'
                AND table_schema = '{schema}'
                AND table_name = '{table}'
            ORDER BY ordinal_position
        """)
        # Handle both dict and array formats
        if columns_data and isinstance(columns_data[0], dict):
            columns = [{"name": row["column_name"], "type": row["data_type"]} for row in columns_data]
        else:
            columns = [{"name": row[0], "type": row[1]} for row in columns_data]
        return {"columns": columns}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching columns: {str(e)}")


@api_app.get("/catalog-tree")
async def get_catalog_tree() -> CatalogTreeResponse:
    """
    DEPRECATED - Use incremental endpoints instead (/catalogs, /schemas, /tables, /columns)
    This loads everything at once and is too slow
    """
    try:
        # Just return catalogs - frontend should use incremental endpoints
        catalogs_data = execute_sql("""
            SELECT catalog_name 
            FROM system.information_schema.catalogs 
            WHERE catalog_owner IS NOT NULL 
            ORDER BY catalog_name
        """)
        # Handle both dict and array formats
        if catalogs_data and isinstance(catalogs_data[0], dict):
            catalogs = [{"name": row["catalog_name"], "schemas": []} for row in catalogs_data]
        else:
            catalogs = [{"name": row[0], "schemas": []} for row in catalogs_data]
        return CatalogTreeResponse(catalogs=catalogs)
    
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error fetching catalog tree: {str(e)}"
        )


# Helper functions for type conversion
def to_float(val, default=0.0):
    """Safely convert value to float"""
    try:
        return float(val) if val is not None and val != '' else default
    except (ValueError, TypeError):
        return default

def to_int(val, default=0):
    """Safely convert value to int"""
    try:
        return int(float(val)) if val is not None and val != '' else default
    except (ValueError, TypeError):
        return default

def to_bool(val):
    """Safely convert value to bool"""
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.lower() in ('true', '1', 'yes')
    return bool(val)

def to_str(val):
    """Safely convert value to string"""
    return str(val) if val is not None else ''


def map_profiling_array_to_dict(row: list, field_name: str) -> dict:
    """
    Map profiling query result array to dictionary with correct column names
    The SQL query structure is defined in utils/sqlProfiler.ts
    Column order is: catalog, schema, table, column, type, [stats based on data type]
    """
    if not row or len(row) < 5:
        return {}
    
    # Detect data type from position 4 (documented_type)
    documented_type = str(row[4]).lower() if len(row) > 4 else ''
    
    # Check data type categories
    is_numeric = any(t in documented_type for t in ['bigint', 'int', 'smallint', 'tinyint', 'decimal', 'double', 'float', 'long'])
    is_date = any(t in documented_type for t in ['date', 'timestamp'])
    is_string = 'string' in documented_type or 'varchar' in documented_type
    
    print(f"DEBUG {field_name}: documented_type='{documented_type}', is_numeric={is_numeric}, is_string={is_string}, row_length={len(row)}")
    
    # Base columns (always present) - positions 0-14
    result = {
        "catalog_name": row[0] if len(row) > 0 else None,
        "schema_name": row[1] if len(row) > 1 else None,
        "table_name": row[2] if len(row) > 2 else None,
        "column_name": row[3] if len(row) > 3 else None,
        "documented_type": row[4] if len(row) > 4 else None,
        "total_rows": to_int(row[5]) if len(row) > 5 else 0,
        "non_null_count": to_int(row[6]) if len(row) > 6 else 0,
        "null_count": to_int(row[7]) if len(row) > 7 else 0,
        "null_percentage": to_float(row[8]) if len(row) > 8 else 0.0,
        "unique_count": to_int(row[9]) if len(row) > 9 else 0,
        "cardinality_pct": to_float(row[10]) if len(row) > 10 else 0.0,
        "unique_percentage": to_float(row[11]) if len(row) > 11 else 0.0,
        "duplicate_percentage": to_float(row[12]) if len(row) > 12 else 0.0,
        "is_categorical": to_bool(row[13]) if len(row) > 13 else False,
        "captured_values_count": to_int(row[14]) if len(row) > 14 else 0,
    }
    
    # Position tracking for conditional fields
    pos = 15
    
    # String-specific fields (avg_length, min_length, max_length, median_length)
    if is_string and len(row) > pos + 3:
        result["avg_length"] = to_float(row[pos])
        result["min_length"] = to_float(row[pos + 1])
        result["max_length"] = to_float(row[pos + 2])
        result["median_length"] = to_float(row[pos + 3])
        pos += 4
    
    # Numeric-specific fields (min, max, mean, stddev, median, p25, p75, p95, p99, zeros, negatives, infinites)
    if is_numeric and len(row) > pos + 11:
        result["min_value"] = to_float(row[pos], None)
        result["max_value"] = to_float(row[pos + 1], None)
        result["mean_value"] = to_float(row[pos + 2], None)
        result["stddev_value"] = to_float(row[pos + 3], None)
        result["median_value"] = to_float(row[pos + 4], None)
        result["p25_value"] = to_float(row[pos + 5], None)
        result["p75_value"] = to_float(row[pos + 6], None)
        result["p95_value"] = to_float(row[pos + 7], None)
        result["p99_value"] = to_float(row[pos + 8], None)
        result["zeros_count"] = to_int(row[pos + 9])
        result["negatives_count"] = to_int(row[pos + 10])
        result["infinite_count"] = to_int(row[pos + 11])
        pos += 12
        print(f"DEBUG {field_name} numeric stats: min={result['min_value']}, max={result['max_value']}, mean={result['mean_value']}, zeros={result['zeros_count']}, negatives={result['negatives_count']}")
    else:
        print(f"DEBUG {field_name}: SKIPPED numeric stats (is_numeric={is_numeric}, row_length={len(row)}, needed={pos + 12})")
    
    # Date-specific fields (min_date, max_date)
    if is_date and len(row) > pos + 1:
        result["min_date"] = row[pos]  # Keep as string
        result["max_date"] = row[pos + 1]  # Keep as string
        pos += 2
    
    # Always present: inferred_type, type_confidence_pct, top_values, all_values
    if len(row) > pos:
        result["inferred_type"] = row[pos]
        pos += 1
    if len(row) > pos:
        result["type_confidence_pct"] = to_float(row[pos])
        pos += 1
    if len(row) > pos:
        # Parse JSON string to array and convert numeric fields
        top_values_str = row[pos]
        try:
            top_values = json.loads(top_values_str) if isinstance(top_values_str, str) else top_values_str
            # Convert frequency and frequency_pct to numbers
            if isinstance(top_values, list):
                for item in top_values:
                    if isinstance(item, dict):
                        if 'frequency' in item:
                            item['frequency'] = to_int(item['frequency'])
                        if 'frequency_pct' in item:
                            item['frequency_pct'] = to_float(item['frequency_pct'])
            result["top_values"] = top_values
        except:
            result["top_values"] = []
        pos += 1
    if len(row) > pos:
        # Parse JSON string to array and convert numeric fields
        all_values_str = row[pos]
        try:
            all_values = json.loads(all_values_str) if isinstance(all_values_str, str) else all_values_str
            # Convert frequency and frequency_pct to numbers
            if isinstance(all_values, list):
                for item in all_values:
                    if isinstance(item, dict):
                        if 'frequency' in item:
                            item['frequency'] = to_int(item['frequency'])
                        if 'frequency_pct' in item:
                            item['frequency_pct'] = to_float(item['frequency_pct'])
            result["all_values"] = all_values
        except:
            result["all_values"] = []
        pos += 1
    
    # patterns (always present now, but may be NULL for non-strings)
    if len(row) > pos:
        patterns_str = row[pos]
        try:
            patterns = json.loads(patterns_str) if isinstance(patterns_str, str) else patterns_str
            # Convert pattern_count and avg_pattern_length to numbers
            if isinstance(patterns, list):
                for item in patterns:
                    if isinstance(item, dict):
                        if 'pattern_count' in item:
                            item['pattern_count'] = to_int(item['pattern_count'])
                        if 'avg_pattern_length' in item:
                            item['avg_pattern_length'] = to_float(item['avg_pattern_length'])
            result["patterns"] = patterns if patterns else []
        except:
            result["patterns"] = []
        pos += 1
    
    # smallest_values (numeric only, but always present as NULL for non-numeric)
    if len(row) > pos:
        smallest_str = row[pos]
        try:
            smallest = json.loads(smallest_str) if isinstance(smallest_str, str) else smallest_str
            if isinstance(smallest, list):
                # Sort by 'rn' field to ensure correct order (smallest first)
                sorted_smallest = sorted([item for item in smallest if isinstance(item, dict) and item.get('value') is not None], 
                                        key=lambda x: x.get('rn', 999))
                result["smallest_values"] = [to_float(item.get('value')) for item in sorted_smallest]
            else:
                result["smallest_values"] = []
        except:
            result["smallest_values"] = []
        pos += 1
    
    # largest_values (numeric only, but always present as NULL for non-numeric)
    if len(row) > pos:
        largest_str = row[pos]
        try:
            largest = json.loads(largest_str) if isinstance(largest_str, str) else largest_str
            if isinstance(largest, list):
                # Sort by 'rn' field to ensure correct order (largest first)
                sorted_largest = sorted([item for item in largest if isinstance(item, dict) and item.get('value') is not None],
                                       key=lambda x: x.get('rn', 999))
                result["largest_values"] = [to_float(item.get('value')) for item in sorted_largest]
            else:
                result["largest_values"] = []
        except:
            result["largest_values"] = []
        pos += 1
    
    # first_samples (always present)
    if len(row) > pos:
        first_samples_str = row[pos]
        try:
            first_samples = json.loads(first_samples_str) if isinstance(first_samples_str, str) else first_samples_str
            if isinstance(first_samples, list):
                # Sort by 'sample_rn' to maintain order
                sorted_first = sorted([item for item in first_samples if isinstance(item, dict) and item.get('sample_value') is not None],
                                     key=lambda x: x.get('sample_rn', 999))
                result["first_samples"] = [str(item.get('sample_value')) for item in sorted_first]
            else:
                result["first_samples"] = []
        except:
            result["first_samples"] = []
        pos += 1
    
    # random_samples (always present)
    if len(row) > pos:
        random_samples_str = row[pos]
        try:
            random_samples = json.loads(random_samples_str) if isinstance(random_samples_str, str) else random_samples_str
            if isinstance(random_samples, list):
                # Sort by 'sample_rn' to maintain order
                sorted_random = sorted([item for item in random_samples if isinstance(item, dict) and item.get('sample_value') is not None],
                                      key=lambda x: x.get('sample_rn', 999))
                result["random_samples"] = [str(item.get('sample_value')) for item in sorted_random]
            else:
                result["random_samples"] = []
        except:
            result["random_samples"] = []
        pos += 1
    
    # Log what was extracted for debugging
    stats_summary = []
    if 'min_value' in result and result['min_value'] is not None:
        stats_summary.append(f"min={result['min_value']}")
    if 'max_value' in result and result['max_value'] is not None:
        stats_summary.append(f"max={result['max_value']}")
    if 'mean_value' in result and result['mean_value'] is not None:
        stats_summary.append(f"mean={result['mean_value']:.2f}")
    
    stats_str = ", ".join(stats_summary) if stats_summary else "no numeric stats"
    print(f"Mapped {field_name} ({documented_type}): {stats_str}")
    
    return result


@api_app.post("/databricks/execute")
async def execute_queries(
    request: QueryRequest
) -> QueryExecutionResponse:
    """
    Execute profiling queries using SQL Statements API
    Handles both dict results (with metadata) and array results (without metadata)
    """
    try:
        results = []
        
        for query_obj in request.queries:
            try:
                # execute_sql returns either list of dicts or list of arrays
                sql_result = execute_sql(query_obj.query)
                
                if sql_result and len(sql_result) > 0:
                    row = sql_result[0]
                    
                    # Check if result is already a dict (has column metadata)
                    if isinstance(row, dict):
                        data_dict = row
                        print(f"Using dict result for {query_obj.fieldKey}: {len(row)} fields")
                    else:
                        # Result is an array - use hardcoded mapping based on SQL structure
                        data_dict = map_profiling_array_to_dict(row, query_obj.fieldKey)
                    
                    results.append({
                        "fieldKey": query_obj.fieldKey,
                        "description": query_obj.description,
                        "data": data_dict,
                        "success": True
                    })
                else:
                    results.append({
                        "fieldKey": query_obj.fieldKey,
                        "description": query_obj.description,
                        "data": None,
                        "success": True
                    })
                
            except Exception as e:
                import traceback
                error_details = traceback.format_exc()
                print(f"Error processing query {query_obj.fieldKey}: {error_details}")
                results.append({
                    "fieldKey": query_obj.fieldKey,
                    "description": query_obj.description,
                    "error": str(e),
                    "success": False
                })
        
        return QueryExecutionResponse(
            results=results,
            success=True,
            message="Queries executed successfully"
        )
    
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Fatal error executing queries: {error_details}")
        raise HTTPException(
            status_code=500,
            detail=f"Error executing queries: {str(e)}"
        )


@api_app.post("/databricks/execute-stream")
async def execute_queries_stream(request: QueryRequest):
    """
    Execute profiling queries with real-time progress updates (streaming)
    Sends progress updates as each query completes using Server-Sent Events
    """
    async def generate_progress():
        try:
            total_queries = len(request.queries)
            results = []
            
            for idx, query_obj in enumerate(request.queries):
                # Send progress update
                progress_data = {
                    "type": "progress",
                    "current": idx + 1,
                    "total": total_queries,
                    "fieldKey": query_obj.fieldKey,
                    "description": query_obj.description,
                    "percentage": round(((idx + 1) / total_queries) * 100, 1)
                }
                yield f"data: {json.dumps(progress_data)}\n\n"
                
                try:
                    # Execute the query
                    sql_result = execute_sql(query_obj.query)
                    
                    if sql_result and len(sql_result) > 0:
                        row = sql_result[0]
                        
                        # Check if result is already a dict (has column metadata)
                        if isinstance(row, dict):
                            data_dict = row
                        else:
                            # Result is an array - use hardcoded mapping
                            data_dict = map_profiling_array_to_dict(row, query_obj.fieldKey)
                        
                        result = {
                            "fieldKey": query_obj.fieldKey,
                            "description": query_obj.description,
                            "data": data_dict,
                            "success": True
                        }
                    else:
                        result = {
                            "fieldKey": query_obj.fieldKey,
                            "description": query_obj.description,
                            "data": None,
                            "success": True
                        }
                    
                    results.append(result)
                    
                    # Send result update
                    result_data = {
                        "type": "result",
                        "result": result
                    }
                    yield f"data: {json.dumps(result_data)}\n\n"
                    
                except Exception as e:
                    import traceback
                    error_details = traceback.format_exc()
                    print(f"Error processing query {query_obj.fieldKey}: {error_details}")
                    
                    error_result = {
                        "fieldKey": query_obj.fieldKey,
                        "description": query_obj.description,
                        "error": str(e),
                        "success": False
                    }
                    results.append(error_result)
                    
                    # Send error update
                    error_data = {
                        "type": "result",
                        "result": error_result
                    }
                    yield f"data: {json.dumps(error_data)}\n\n"
            
            # Send completion message
            complete_data = {
                "type": "complete",
                "results": results,
                "total": len(results)
            }
            yield f"data: {json.dumps(complete_data)}\n\n"
            
        except Exception as e:
            import traceback
            print(f"Error in execute_queries_stream: {traceback.format_exc()}")
            error_data = {
                "type": "error",
                "error": str(e)
            }
            yield f"data: {json.dumps(error_data)}\n\n"
    
    return StreamingResponse(
        generate_progress(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )


@api_app.get("/catalogs")
async def list_catalogs():
    """DEPRECATED - Use /api/catalogs instead. List all available catalogs"""
    try:
        data = execute_sql("SHOW CATALOGS")
        # SHOW CATALOGS returns: catalog
        catalogs = [row.get("catalog", list(row.values())[0]) if isinstance(row, dict) else str(row) for row in data]
        return {"catalogs": catalogs}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_app.get("/schemas/{catalog}")
async def list_schemas(catalog: str):
    """DEPRECATED - Use /api/schemas?catalog=X instead. List schemas in a catalog"""
    try:
        data = execute_sql(f"SHOW SCHEMAS IN `{catalog}`")
        # SHOW SCHEMAS returns: databaseName
        schemas = [row.get("databaseName", list(row.values())[0]) if isinstance(row, dict) else str(row) for row in data if (row.get("databaseName") if isinstance(row, dict) else row) != "information_schema"]
        return {"schemas": schemas}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_app.get("/tables/{catalog}/{schema}")
async def list_tables(catalog: str, schema: str):
    """DEPRECATED - Use /api/tables?catalog=X&schema=Y instead. List tables in a schema"""
    try:
        data = execute_sql(f"SHOW TABLES IN `{catalog}`.`{schema}`")
        # SHOW TABLES returns: database, tableName, isTemporary
        tables = [row.get("tableName", list(row.values())[1] if len(row.values()) > 1 else list(row.values())[0]) if isinstance(row, dict) else str(row) for row in data]
        return {"tables": tables}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# UI Application (Serves React Frontend)
# ============================================================================

ui_app = FastAPI()

# Mount API under /api prefix
# This ensures all API calls go to /api/* and UI handles the rest
ui_app.mount("/api", api_app)

@api_app.post("/export/excel")
async def export_to_excel(profile_data: Dict[str, Any]):
    """
    Export profiling results to Excel file with multiple sheets
    """
    try:
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Define header styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Sheet 1: Summary
        ws_summary = wb.create_sheet("Summary")
        summary_data = [
            ["Profile Summary", ""],
            ["Total Columns", profile_data.get("totalColumns", 0)],
            ["Total Rows", profile_data.get("totalRows", 0)],
            ["Issues Found", profile_data.get("issuesFound", 0)],
            ["Completeness %", profile_data.get("completeness", "0%")],
            ["Quality Score", profile_data.get("qualityScore", "0%")],
            ["High Cardinality Count", profile_data.get("highCardinalityCount", 0)],
            ["Date Columns", profile_data.get("dateColumns", 0)],
            ["Empty Columns", profile_data.get("emptyColumns", 0)],
        ]
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=cell_value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
        
        # Sheet 2: Detailed Column Data
        ws_detailed = wb.create_sheet("Detailed Profiling")
        detailed_headers = [
            "Catalog", "Schema", "Table", "Column Name", "Data Type", "Inferred Type",
            "Unique Values", "Unique %", "Nulls", "Null %", "Completeness %",
            "Min Value", "Max Value", "Mean", "Median", "Std Dev",
            "Min Length", "Max Length", "Avg Length",
            "Zeros", "Negatives", "Infinites",
            "Quality Score"
        ]
        
        for col_idx, header in enumerate(detailed_headers, 1):
            cell = ws_detailed.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        columns = profile_data.get("columns", [])
        for row_idx, col_data in enumerate(columns, 2):
            ws_detailed.cell(row=row_idx, column=1, value=col_data.get("catalog", ""))
            ws_detailed.cell(row=row_idx, column=2, value=col_data.get("schema", ""))
            ws_detailed.cell(row=row_idx, column=3, value=col_data.get("table", ""))
            ws_detailed.cell(row=row_idx, column=4, value=col_data.get("name", ""))
            ws_detailed.cell(row=row_idx, column=5, value=col_data.get("documentedType", ""))
            ws_detailed.cell(row=row_idx, column=6, value=col_data.get("inferredType", ""))
            ws_detailed.cell(row=row_idx, column=7, value=col_data.get("uniqueValues", 0))
            ws_detailed.cell(row=row_idx, column=8, value=col_data.get("uniquePct", 0))
            ws_detailed.cell(row=row_idx, column=9, value=col_data.get("nulls", 0))
            ws_detailed.cell(row=row_idx, column=10, value=col_data.get("nullPct", 0))
            ws_detailed.cell(row=row_idx, column=11, value=100 - col_data.get("nullPct", 0))
            ws_detailed.cell(row=row_idx, column=12, value=col_data.get("minValue", ""))
            ws_detailed.cell(row=row_idx, column=13, value=col_data.get("maxValue", ""))
            ws_detailed.cell(row=row_idx, column=14, value=col_data.get("mean"))
            ws_detailed.cell(row=row_idx, column=15, value=col_data.get("median"))
            ws_detailed.cell(row=row_idx, column=16, value=col_data.get("stddev"))
            ws_detailed.cell(row=row_idx, column=17, value=col_data.get("minLength"))
            ws_detailed.cell(row=row_idx, column=18, value=col_data.get("maxLength"))
            ws_detailed.cell(row=row_idx, column=19, value=col_data.get("avgLength"))
            ws_detailed.cell(row=row_idx, column=20, value=col_data.get("zerosCount"))
            ws_detailed.cell(row=row_idx, column=21, value=col_data.get("negativesCount"))
            ws_detailed.cell(row=row_idx, column=22, value=col_data.get("infiniteCount"))
            # Calculate quality score (simplified)
            quality = 100 - min(col_data.get("nullPct", 0), 30)
            ws_detailed.cell(row=row_idx, column=23, value=quality)
        
        # Sheet 3: Samples
        ws_samples = wb.create_sheet("Sample Values")
        sample_headers = ["Column Name", "Sample Type", "Sample 1", "Sample 2", "Sample 3", 
                         "Sample 4", "Sample 5", "Sample 6", "Sample 7", "Sample 8", "Sample 9", "Sample 10"]
        
        for col_idx, header in enumerate(sample_headers, 1):
            cell = ws_samples.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        sample_row = 2
        for col_data in columns:
            # First samples
            first_samples = col_data.get("firstSamples", [])
            if first_samples:
                ws_samples.cell(row=sample_row, column=1, value=col_data.get("name", ""))
                ws_samples.cell(row=sample_row, column=2, value="First 10")
                for idx, sample in enumerate(first_samples[:10], 3):
                    ws_samples.cell(row=sample_row, column=idx, value=str(sample))
                sample_row += 1
            
            # Random samples
            random_samples = col_data.get("randomSamples", [])
            if random_samples:
                ws_samples.cell(row=sample_row, column=1, value=col_data.get("name", ""))
                ws_samples.cell(row=sample_row, column=2, value="Random 10")
                for idx, sample in enumerate(random_samples[:10], 3):
                    ws_samples.cell(row=sample_row, column=idx, value=str(sample))
                sample_row += 1
        
        # Sheet 4: Extreme Values
        ws_extremes = wb.create_sheet("Extreme Values")
        extreme_headers = ["Column Name", "Value Type", "Value 1", "Value 2", "Value 3", "Value 4", "Value 5"]
        
        for col_idx, header in enumerate(extreme_headers, 1):
            cell = ws_extremes.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        extreme_row = 2
        for col_data in columns:
            # Smallest values
            smallest = col_data.get("smallestValues", [])
            if smallest:
                ws_extremes.cell(row=extreme_row, column=1, value=col_data.get("name", ""))
                ws_extremes.cell(row=extreme_row, column=2, value="Smallest")
                for idx, val in enumerate(smallest[:5], 3):
                    ws_extremes.cell(row=extreme_row, column=idx, value=val)
                extreme_row += 1
            
            # Largest values
            largest = col_data.get("largestValues", [])
            if largest:
                ws_extremes.cell(row=extreme_row, column=1, value=col_data.get("name", ""))
                ws_extremes.cell(row=extreme_row, column=2, value="Largest")
                for idx, val in enumerate(largest[:5], 3):
                    ws_extremes.cell(row=extreme_row, column=idx, value=val)
                extreme_row += 1
        
        # Auto-size columns for all sheets
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=data_profiling_report.xlsx"}
        )
    
    except Exception as e:
        import traceback
        print(f"Error creating Excel export: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error creating Excel export: {str(e)}")


# ============================================================================
# Cross-Column Analysis Endpoints
# ============================================================================

@api_app.post("/cross-column/correlations")
async def calculate_correlations(request: dict):
    """
    Calculate correlation matrix for numeric columns.
    """
    try:
        catalog = request.get("catalog")
        schema = request.get("schema")
        table = request.get("table")
        numeric_fields = request.get("numericFields", [])
        
        if len(numeric_fields) < 2:
            return {"correlations": [], "message": "Need at least 2 numeric columns"}
        
        # Build correlation SQL
        correlations = []
        
        for i, field1 in enumerate(numeric_fields):
            for j, field2 in enumerate(numeric_fields):
                if i < j:  # Only calculate upper triangle
                    sql = f"""
                    SELECT 
                        CORR(CAST(`{field1}` AS DOUBLE), CAST(`{field2}` AS DOUBLE)) as correlation
                    FROM `{catalog}`.`{schema}`.`{table}`
                    WHERE `{field1}` IS NOT NULL AND `{field2}` IS NOT NULL
                    """
                    
                    result = execute_sql(sql)
                    
                    if result and len(result) > 0:
                        row = result[0]
                        corr_value = to_float(row.get('correlation') if isinstance(row, dict) else row[0])
                        
                        correlations.append({
                            "field1": field1,
                            "field2": field2,
                            "correlation": corr_value if corr_value is not None else 0.0,
                            "strength": get_correlation_strength(corr_value if corr_value else 0.0)
                        })
        
        return {"correlations": correlations}
    
    except Exception as e:
        import traceback
        print(f"Error calculating correlations: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


def get_correlation_strength(corr: float) -> str:
    """Classify correlation strength."""
    abs_corr = abs(corr)
    if abs_corr >= 0.8:
        return "Very Strong"
    elif abs_corr >= 0.6:
        return "Strong"
    elif abs_corr >= 0.4:
        return "Moderate"
    elif abs_corr >= 0.2:
        return "Weak"
    else:
        return "Very Weak"


@api_app.post("/cross-column/composite-keys")
async def detect_composite_keys(request: dict):
    """
    Detect potential composite key candidates.
    """
    try:
        catalog = request.get("catalog")
        schema = request.get("schema")
        table = request.get("table")
        fields = request.get("fields", [])
        
        if len(fields) < 2:
            return {"compositeKeys": [], "message": "Need at least 2 columns"}
        
        # Get total row count
        count_sql = f"SELECT COUNT(*) as total FROM `{catalog}`.`{schema}`.`{table}`"
        count_result = execute_sql(count_sql)
        total_rows = to_int(count_result[0].get('total') if isinstance(count_result[0], dict) else count_result[0][0])
        
        composite_keys = []
        
        # Check 2-column combinations
        for i, field1 in enumerate(fields):
            for j, field2 in enumerate(fields):
                if i < j:
                    sql = f"""
                    SELECT COUNT(DISTINCT `{field1}`, `{field2}`) as unique_count
                    FROM `{catalog}`.`{schema}`.`{table}`
                    """
                    
                    result = execute_sql(sql)
                    unique_count = to_int(result[0].get('unique_count') if isinstance(result[0], dict) else result[0][0])
                    uniqueness_pct = (unique_count / total_rows * 100) if total_rows > 0 else 0
                    
                    if uniqueness_pct >= 95:  # Potential key if >= 95% unique
                        composite_keys.append({
                            "columns": [field1, field2],
                            "uniqueCount": unique_count,
                            "totalRows": total_rows,
                            "uniquenessPct": round(uniqueness_pct, 2),
                            "isPotentialKey": uniqueness_pct >= 99.9
                        })
        
        # Sort by uniqueness descending
        composite_keys.sort(key=lambda x: x["uniquenessPct"], reverse=True)
        
        return {"compositeKeys": composite_keys[:10]}  # Top 10
    
    except Exception as e:
        import traceback
        print(f"Error detecting composite keys: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@api_app.post("/cross-column/conditional-profiling")
async def conditional_profiling(request: dict):
    """
    Profile numeric columns grouped by categorical columns.
    """
    try:
        catalog = request.get("catalog")
        schema = request.get("schema")
        table = request.get("table")
        numeric_field = request.get("numericField")
        categorical_field = request.get("categoricalField")
        
        sql = f"""
        WITH stats AS (
            SELECT 
                `{categorical_field}` as category,
                COUNT(*) as count,
                AVG(CAST(`{numeric_field}` AS DOUBLE)) as mean_value,
                STDDEV(CAST(`{numeric_field}` AS DOUBLE)) as stddev_value,
                MIN(CAST(`{numeric_field}` AS DOUBLE)) as min_value,
                MAX(CAST(`{numeric_field}` AS DOUBLE)) as max_value,
                PERCENTILE(CAST(`{numeric_field}` AS DOUBLE), 0.5) as median_value
            FROM `{catalog}`.`{schema}`.`{table}`
            WHERE `{categorical_field}` IS NOT NULL 
              AND `{numeric_field}` IS NOT NULL
            GROUP BY `{categorical_field}`
            ORDER BY count DESC
            LIMIT 20
        )
        SELECT * FROM stats
        """
        
        result = execute_sql(sql)
        
        conditional_stats = []
        for row in result:
            if isinstance(row, dict):
                conditional_stats.append({
                    "category": str(row.get('category', '')),
                    "count": to_int(row.get('count', 0)),
                    "mean": to_float(row.get('mean_value', 0)),
                    "stddev": to_float(row.get('stddev_value', 0)),
                    "min": to_float(row.get('min_value', 0)),
                    "max": to_float(row.get('max_value', 0)),
                    "median": to_float(row.get('median_value', 0))
                })
            else:
                conditional_stats.append({
                    "category": str(row[0]),
                    "count": to_int(row[1]),
                    "mean": to_float(row[2]),
                    "stddev": to_float(row[3]),
                    "min": to_float(row[4]),
                    "max": to_float(row[5]),
                    "median": to_float(row[6])
                })
        
        return {"conditionalStats": conditional_stats}
    
    except Exception as e:
        import traceback
        print(f"Error in conditional profiling: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# Temporal Analysis Endpoint
# ============================================================================

@api_app.post("/temporal/analyze")
async def analyze_temporal_column(request: dict):
    """
    Analyze temporal patterns in a timestamp/date column.
    Returns day-of-week and hour-of-day distributions.
    """
    try:
        catalog = request.get("catalog")
        schema = request.get("schema")
        table = request.get("table")
        column = request.get("column")
        
        # Day of week analysis
        dow_sql = f"""
        WITH day_analysis AS (
            SELECT 
                DAYOFWEEK(`{column}`) as day_num,
                CASE DAYOFWEEK(`{column}`)
                    WHEN 1 THEN 'Sunday'
                    WHEN 2 THEN 'Monday'
                    WHEN 3 THEN 'Tuesday'
                    WHEN 4 THEN 'Wednesday'
                    WHEN 5 THEN 'Thursday'
                    WHEN 6 THEN 'Friday'
                    WHEN 7 THEN 'Saturday'
                END as day_name,
                COUNT(*) as count
            FROM `{catalog}`.`{schema}`.`{table}`
            WHERE `{column}` IS NOT NULL
            GROUP BY DAYOFWEEK(`{column}`)
            ORDER BY DAYOFWEEK(`{column}`)
        )
        SELECT day_name, count FROM day_analysis
        """
        
        dow_result = execute_sql(dow_sql)
        day_of_week = []
        for row in dow_result:
            if isinstance(row, dict):
                day_of_week.append({
                    "day": row.get('day_name', ''),
                    "count": to_int(row.get('count', 0))
                })
            else:
                day_of_week.append({
                    "day": str(row[0]),
                    "count": to_int(row[1])
                })
        
        # Hour of day analysis
        hod_sql = f"""
        WITH hour_analysis AS (
            SELECT 
                HOUR(`{column}`) as hour,
                COUNT(*) as count
            FROM `{catalog}`.`{schema}`.`{table}`
            WHERE `{column}` IS NOT NULL
            GROUP BY HOUR(`{column}`)
            ORDER BY HOUR(`{column}`)
        )
        SELECT hour, count FROM hour_analysis
        """
        
        hod_result = execute_sql(hod_sql)
        hour_of_day = []
        for row in hod_result:
            if isinstance(row, dict):
                hour_of_day.append({
                    "hour": to_int(row.get('hour', 0)),
                    "count": to_int(row.get('count', 0))
                })
            else:
                hour_of_day.append({
                    "hour": to_int(row[0]),
                    "count": to_int(row[1])
                })
        
        return {
            "dayOfWeek": day_of_week,
            "hourOfDay": hour_of_day
        }
    
    except Exception as e:
        import traceback
        print(f"Error analyzing temporal column: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# AI-Powered Insights Endpoints
# ============================================================================

@api_app.post("/ai/generate-insights")
async def generate_ai_insights(request: dict):
    """
    Generate AI-powered insights for a column using Databricks Foundation Models.
    """
    try:
        column_data = request.get("columnData")
        
        # Build prompt
        prompt = f"""You are a data quality expert. Analyze this column profile and provide 3-5 actionable insights.

Column: {column_data.get('name')}
Data Type: {column_data.get('inferredType')} (documented: {column_data.get('documentedType')})
Total Rows: {column_data.get('totalRows', 0):,}
Null Values: {column_data.get('nulls', 0):,} ({column_data.get('nullPct', 0):.2f}%)
Unique Values: {column_data.get('uniqueValues', 0):,} ({column_data.get('uniquePct', 0):.2f}%)
Completeness: {column_data.get('completeness', 0):.2f}%
Quality Score: {column_data.get('quality', 0)}/100

"""
        
        # Add numeric stats if available
        if column_data.get('mean') is not None:
            prompt += f"""Statistical Summary:
- Mean: {column_data.get('mean', 0):.2f}
- Median: {column_data.get('median', 0):.2f}
- Std Dev: {column_data.get('stddev', 0):.2f}
- Min: {column_data.get('minValue', 0)}
- Max: {column_data.get('maxValue', 0)}
- Zeros: {column_data.get('zerosCount', 0):,}
- Negatives: {column_data.get('negativesCount', 0):,}

"""
        
        # Add top values
        if column_data.get('topValues'):
            prompt += "Top Values:\n"
            for val in column_data.get('topValues', [])[:5]:
                prompt += f"- {val.get('value')}: {val.get('count', 0):,} occurrences ({val.get('percentage', 0):.2f}%)\n"
        
        prompt += """
Provide insights in this format:
1. [Icon] [Title]: [Description]
2. [Icon] [Title]: [Description]
...

Use these icons: ⚠️ (warning), 💡 (suggestion), ✅ (good practice), 🔍 (observation), 📊 (data pattern)

Focus on:
- Data quality issues and how to fix them
- Unusual patterns or anomalies
- Best practices being followed or violated
- Recommendations for data cleaning or validation
- Business insights from the distribution
"""
        
        # Call Databricks Foundation Models API
        try:
            token = get_oauth_token()
            endpoint_url = f"https://{os.getenv('DATABRICKS_SERVER_HOSTNAME')}/serving-endpoints/databricks-gemma-3-12b/invocations"
            
            headers = {
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json"
            }
            
            payload = {
                "messages": [
                    {
                        "role": "user",
                        "content": prompt
                    }
                ],
                "max_tokens": 1000,
                "temperature": 0.7
            }
            
            response = requests.post(endpoint_url, json=payload, headers=headers, timeout=30)
            
            if response.status_code == 200:
                result = response.json()
                insights_text = result.get('choices', [{}])[0].get('message', {}).get('content', '')
                
                return {
                    "insights": insights_text,
                    "success": True
                }
            else:
                # Fallback to rule-based insights if AI fails
                return {
                    "insights": generate_rule_based_insights(column_data),
                    "success": True,
                    "fallback": True
                }
        
        except Exception as ai_error:
            print(f"AI API error: {ai_error}, falling back to rule-based insights")
            return {
                "insights": generate_rule_based_insights(column_data),
                "success": True,
                "fallback": True
            }
    
    except Exception as e:
        import traceback
        print(f"Error generating AI insights: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


def generate_rule_based_insights(column_data: dict) -> str:
    """Generate rule-based insights as fallback."""
    insights = []
    
    # Null check
    null_pct = column_data.get('nullPct', 0)
    if null_pct == 0:
        insights.append("✅ **Perfect Completeness**: Column has no null values, indicating reliable data capture.")
    elif null_pct < 5:
        insights.append(f"⚠️ **Minor Nulls**: {null_pct:.2f}% null values detected. Consider if nulls are expected or indicate data quality issues.")
    elif null_pct < 20:
        insights.append(f"⚠️ **Moderate Nulls**: {null_pct:.2f}% null values present. Investigate root cause and consider imputation strategies.")
    else:
        insights.append(f"🔴 **High Nulls**: {null_pct:.2f}% null values is significant. This column may have data collection issues.")
    
    # Uniqueness check
    unique_pct = column_data.get('uniquePct', 0)
    if unique_pct > 99:
        insights.append("🔑 **Potential Primary Key**: Nearly 100% unique values. Consider using as a primary key.")
    elif unique_pct < 1 and column_data.get('uniqueValues', 0) < 10:
        insights.append(f"📊 **Low Cardinality**: Only {column_data.get('uniqueValues', 0)} distinct values. Good candidate for categorical analysis or indexing.")
    
    # Numeric insights
    if column_data.get('mean') is not None:
        zeros = column_data.get('zerosCount', 0)
        negatives = column_data.get('negativesCount', 0)
        
        if negatives > 0:
            insights.append(f"🔍 **Negative Values**: Found {negatives:,} negative values. Verify if negative values are valid for this field.")
        
        if zeros > column_data.get('totalRows', 1) * 0.1:
            insights.append(f"⚠️ **Many Zeros**: {zeros:,} zero values ({zeros/column_data.get('totalRows', 1)*100:.1f}%). Check if zeros represent missing data.")
        
        # Check for outliers
        stddev_val = to_float(column_data.get('stddev', 0))
        if stddev_val and stddev_val > 0:
            mean_val = to_float(column_data.get('mean', 0))
            max_val = to_float(column_data.get('maxValue', 0))
            
            if mean_val is not None and max_val is not None and stddev_val is not None:
                if max_val > mean_val + (3 * stddev_val):
                    insights.append("📊 **Outliers Detected**: Maximum value is >3σ from mean. Review extreme values for data entry errors.")
    
    # Quality score
    quality = column_data.get('quality', 0)
    if quality >= 95:
        insights.append(f"✅ **Excellent Quality**: Quality score of {quality}/100 indicates well-maintained data.")
    elif quality < 70:
        insights.append(f"⚠️ **Quality Concerns**: Quality score of {quality}/100. Focus on improving completeness and consistency.")
    
    # If no insights generated, add a generic one
    if not insights:
        insights.append("💡 **Data Looks Reasonable**: No major quality issues detected in this column.")
    
    return "\n\n".join(insights)


# ============================================================================
# Profile Snapshot Management Endpoints
# ============================================================================

# Shared storage for snapshots using filesystem (works across gunicorn workers)
# In production, use a database or Redis for better persistence
profile_snapshots = {}
SNAPSHOTS_DIR = "/tmp/databricks_profiler_snapshots"

def ensure_snapshots_dir():
    """Ensure the snapshots directory exists"""
    os.makedirs(SNAPSHOTS_DIR, exist_ok=True)

def load_snapshots_from_disk():
    """Load all snapshots from disk into memory"""
    ensure_snapshots_dir()
    loaded_count = 0
    for filename in os.listdir(SNAPSHOTS_DIR):
        if filename.endswith('.json'):
            filepath = os.path.join(SNAPSHOTS_DIR, filename)
            try:
                with open(filepath, 'r') as f:
                    snapshot = json.load(f)
                    profile_snapshots[snapshot['id']] = snapshot
                    loaded_count += 1
            except Exception as e:
                print(f"⚠️ Error loading snapshot {filename}: {e}")
    return loaded_count

def save_snapshot_to_disk(snapshot_id: str, snapshot_data: dict):
    """Save a snapshot to disk for persistence across workers"""
    ensure_snapshots_dir()
    filepath = os.path.join(SNAPSHOTS_DIR, f"{snapshot_id}.json")
    with open(filepath, 'w') as f:
        json.dump(snapshot_data, f)
    print(f"💾 Persisted snapshot to disk: {filepath}")

def delete_snapshot_from_disk(snapshot_id: str):
    """Delete a snapshot from disk"""
    filepath = os.path.join(SNAPSHOTS_DIR, f"{snapshot_id}.json")
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"🗑️ Deleted snapshot from disk: {filepath}")

# Load snapshots at startup (for this worker process)
startup_loaded = load_snapshots_from_disk()
print(f"🔄 Worker startup: Loaded {startup_loaded} existing snapshots from {SNAPSHOTS_DIR}")

@api_app.post("/snapshots/save")
async def save_snapshot(request: dict):
    """
    Save a profile snapshot for later comparison.
    """
    try:
        snapshot_name = request.get("name")
        profile_data = request.get("profileData")
        timestamp = request.get("timestamp")
        
        # Generate a unique ID using UUID to prevent collisions
        snapshot_id = str(uuid.uuid4())
        
        # Create a deep copy of the profile data to prevent reference issues
        import copy
        profile_data_copy = copy.deepcopy(profile_data)
        
        snapshot_data = {
            "id": snapshot_id,
            "name": snapshot_name,
            "timestamp": timestamp,
            "data": profile_data_copy
        }
        
        # Save to memory AND disk for persistence across workers
        profile_snapshots[snapshot_id] = snapshot_data
        save_snapshot_to_disk(snapshot_id, snapshot_data)
        
        print(f"✅ Saved snapshot: ID={snapshot_id}, Name={snapshot_name}")
        print(f"   Total snapshots in THIS worker's memory: {len(profile_snapshots)}")
        print(f"   All snapshot IDs in THIS worker: {list(profile_snapshots.keys())}")
        print(f"   All snapshot names in THIS worker: {[s['name'] for s in profile_snapshots.values()]}")
        
        return {
            "success": True,
            "snapshotId": snapshot_id,
            "message": f"Snapshot '{snapshot_name}' saved successfully"
        }
    
    except Exception as e:
        import traceback
        print(f"Error saving snapshot: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@api_app.get("/snapshots/list")
async def list_snapshots():
    """
    List all saved snapshots (reloads from disk to get snapshots from all workers).
    """
    try:
        print(f"📋 Listing snapshots request received")
        
        # Reload from disk to get snapshots saved by other workers
        load_snapshots_from_disk()
        
        print(f"   Total in memory after reload: {len(profile_snapshots)}")
        print(f"   Snapshot IDs: {list(profile_snapshots.keys())}")
        
        snapshots = [
            {
                "id": snap["id"],
                "name": snap["name"],
                "timestamp": snap["timestamp"],
                "columnCount": len(snap["data"].get("columns", []))
            }
            for snap in profile_snapshots.values()
        ]
        
        # Sort by timestamp descending
        snapshots.sort(key=lambda x: x["timestamp"], reverse=True)
        
        print(f"   Returning {len(snapshots)} snapshots to frontend")
        for snap in snapshots:
            print(f"   - {snap['name']} (ID: {snap['id'][:8]}..., {snap['columnCount']} cols)")
        
        return {"snapshots": snapshots}
    
    except Exception as e:
        import traceback
        print(f"Error listing snapshots: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@api_app.get("/snapshots/{snapshot_id}")
async def get_snapshot(snapshot_id: str):
    """
    Retrieve a specific snapshot (reloads from disk if not in memory).
    """
    try:
        # Try memory first, then reload from disk
        if snapshot_id not in profile_snapshots:
            load_snapshots_from_disk()
        
        if snapshot_id not in profile_snapshots:
            raise HTTPException(status_code=404, detail="Snapshot not found")
        
        return {"snapshot": profile_snapshots[snapshot_id]}
    
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"Error retrieving snapshot: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@api_app.delete("/snapshots/{snapshot_id}")
async def delete_snapshot(snapshot_id: str):
    """
    Delete a specific snapshot from memory and disk.
    """
    try:
        # Reload from disk first
        load_snapshots_from_disk()
        
        if snapshot_id not in profile_snapshots:
            raise HTTPException(status_code=404, detail="Snapshot not found")
        
        # Get name for logging before deleting
        snapshot_name = profile_snapshots[snapshot_id].get("name", "Unknown")
        
        # Delete from both memory and disk
        del profile_snapshots[snapshot_id]
        delete_snapshot_from_disk(snapshot_id)
        
        print(f"✅ Deleted snapshot: ID={snapshot_id}, Name={snapshot_name}")
        print(f"   Remaining snapshots in THIS worker: {len(profile_snapshots)}")
        
        return {
            "success": True,
            "message": f"Snapshot {snapshot_id} deleted successfully"
        }
    
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"Error deleting snapshot: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@api_app.post("/snapshots/compare")
async def compare_snapshots(request: dict):
    """
    Compare two profile snapshots (reloads from disk to ensure both are available).
    """
    try:
        snapshot_id_1 = request.get("snapshotId1")
        snapshot_id_2 = request.get("snapshotId2")
        
        print(f"🔍 Comparing snapshots: ID1={snapshot_id_1}, ID2={snapshot_id_2}")
        
        # Reload from disk to get snapshots from all workers
        load_snapshots_from_disk()
        
        print(f"   Available snapshots after reload: {list(profile_snapshots.keys())}")
        
        if snapshot_id_1 not in profile_snapshots or snapshot_id_2 not in profile_snapshots:
            print(f"   ❌ ERROR: Snapshot not found. ID1 exists: {snapshot_id_1 in profile_snapshots}, ID2 exists: {snapshot_id_2 in profile_snapshots}")
            raise HTTPException(status_code=404, detail=f"One or both snapshots not found. Available: {list(profile_snapshots.keys())}")
        
        snap1 = profile_snapshots[snapshot_id_1]
        snap2 = profile_snapshots[snapshot_id_2]
        
        # Compare columns
        comparisons = []
        
        snap1_cols = {col['name']: col for col in snap1['data'].get('columns', [])}
        snap2_cols = {col['name']: col for col in snap2['data'].get('columns', [])}
        
        all_col_names = set(snap1_cols.keys()) | set(snap2_cols.keys())
        
        for col_name in all_col_names:
            col1 = snap1_cols.get(col_name)
            col2 = snap2_cols.get(col_name)
            
            if col1 and col2:
                # Both exist - compare metrics
                comparison = {
                    "columnName": col_name,
                    "status": "changed",
                    "changes": {
                        "nulls": {
                            "before": col1.get('nullPct', 0),
                            "after": col2.get('nullPct', 0),
                            "delta": col2.get('nullPct', 0) - col1.get('nullPct', 0)
                        },
                        "uniqueness": {
                            "before": col1.get('uniquePct', 0),
                            "after": col2.get('uniquePct', 0),
                            "delta": col2.get('uniquePct', 0) - col1.get('uniquePct', 0)
                        },
                        "quality": {
                            "before": col1.get('quality', 0),
                            "after": col2.get('quality', 0),
                            "delta": col2.get('quality', 0) - col1.get('quality', 0)
                        }
                    }
                }
                
                # Add numeric changes if available
                if col1.get('mean') is not None and col2.get('mean') is not None:
                    comparison['changes']['mean'] = {
                        "before": col1.get('mean', 0),
                        "after": col2.get('mean', 0),
                        "delta": col2.get('mean', 0) - col1.get('mean', 0)
                    }
                
                comparisons.append(comparison)
            
            elif col1 and not col2:
                comparisons.append({
                    "columnName": col_name,
                    "status": "removed"
                })
            
            elif col2 and not col1:
                comparisons.append({
                    "columnName": col_name,
                    "status": "added"
                })
        
        return {
            "comparison": {
                "snapshot1": {"name": snap1['name'], "timestamp": snap1['timestamp']},
                "snapshot2": {"name": snap2['name'], "timestamp": snap2['timestamp']},
                "columns": comparisons
            }
        }
    
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"Error comparing snapshots: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


# Mount React static files at root
# The 'html=True' parameter ensures index.html is served for all routes (SPA routing)
ui_app.mount(
    "/",
    StaticFiles(directory="client/build", html=True),
    name="ui"
)


# ============================================================================
# Main Application Export
# ============================================================================

# This is what gunicorn will run
app = ui_app


if __name__ == "__main__":
    # For local development only
    import uvicorn
    print("🚀 Starting Databricks Data Profiler App")
    print("📊 Dashboard: http://localhost:8000")
    print("🔧 API Docs: http://localhost:8000/api/docs")
    uvicorn.run(app, host="0.0.0.0", port=8000)

